import calendar
import os
from datetime import date, timedelta, datetime
import numpy as np
import openpyxl
import pandas as pd
from django.db.models import Sum, Q
from izouapp.models import orders, DeliveryPerson, Pizza
from izouaproject import settings
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib
matplotlib.use('Agg')
import seaborn as sns

img_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'izouapp', 'images','img_gen_from_charts')# chemin où seront stocké les fichiers img


def prepare_datas_to_export(first_period=None,second_period=None):
    fetched_datas = orders.objects.filter(create_at__gte=first_period, create_at__lte=second_period)
    datas_cleaned = []
    if len(fetched_datas):
        for data in fetched_datas:
            date_formatee_fr = data.create_at.strftime("%d %B %Y")
            time_formatee = data.deliveryHour.strftime(
                "%Hh%M") if data.deliveryHour else 'Heure sur place: ' + data.onSiteHour.strftime('%H:%M')
            pizza_names = ", ".join([pizza.__str__() for pizza in data.pizzas.all()])
            surplace = 'Oui' if data.surplace else 'Non'
            delivered = {'delivered': 'Livré', 'canceled': 'Annulé', 'on-site': 'Sur place', 'pending': 'En attente'}[
                data.status]
            methode_payment_order = 'Chez Izoua' if data.payment_method_order == 'izoua' else 'Chez le livreur'
            methode_payment_delivery = 'Chez Izoua' if data.payment_method_delivery == 'izoua' else 'Chez le livreur'
            data_row = [date_formatee_fr, surplace,
                        data.deliveryAdress if data.deliveryAdress else 'Commandée sur place', time_formatee,
                        data.deliveryPerson.name if data.deliveryPerson else 'Commandée sur place',
                        delivered, data.client.name, methode_payment_order if data.surplace else 'Chez Izoua', methode_payment_delivery if data.surplace else 'Chez Izoua', pizza_names,
                        data.deliveryPrice if data.surplace else 'Commandée sur place',
                        data.pizza_and_extratopping_price, data.total_price]

            datas_cleaned.append(data_row)

        return datas_cleaned

    return None


def get_periodicaly_total_orders() -> dict:
    today = date.today()

    # Calcul des périodes
    previous_period_start = today - timedelta(days=today.weekday() + 7)
    previous_period_end = previous_period_start + timedelta(days=6)

    period_before_start = previous_period_start - timedelta(days=7)
    period_before_end = period_before_start + timedelta(days=6)

    # Commandes de la periode dernière
    previous_period_orders = orders.objects.filter(
        create_at__gte=previous_period_start, create_at__lte=previous_period_end
    )
    previous_period_order_count = previous_period_orders.count()

    # Commandes de la periode d'avant
    period_before_orders = orders.objects.filter(
        create_at__gte=period_before_start, create_at__lte=period_before_end,
    )
    period_before_order_count = period_before_orders.count()

    return {'previous_period_order_count':previous_period_order_count,'period_before_order_count':period_before_order_count}


def get_periodicaly_orders_info(filter_conditions=None, period="week") -> dict|bool:
    today = date.today()

    if period == "week":
        # Calcul des périodes
        previous_period_start = today - timedelta(days=today.weekday() + 7)
        previous_period_end = previous_period_start + timedelta(days=6)

        before_period_start = previous_period_start - timedelta(days=7)
        before_period_end = before_period_start + timedelta(days=6)

    else:
        month1 = 12 if today.month - 1 == 0 else today.month - 1 # le mois précedent 1,2,...,12
        month2 = month1 - 1 # le mois d'avant 1,2,...,12
        year = today.year-1 if today.month - 1 == 0 else today.year # l'année actuelle

        # Obtenir le premier jour du mois précédent
        previous_period_start = datetime(year, month1, 1)

        # Obtenir le dernier jour du mois précédent
        previous_period_end = datetime(year, month1,
                                           calendar.monthrange(year, month1)[1])

        # Obtenir le premier jour du mois d'avant
        before_period_start = datetime(year, month2, 1)

        # Obtenir le dernier jour du mois d'avant
        before_period_end = datetime(year, month2,
                                           calendar.monthrange(year, month2)[1])

    # Filtre optionnel
    filter_conditions = filter_conditions or {}

    # Commandes de la periode dernière
    previous_period_orders = orders.objects.filter(create_at__gte=previous_period_start, create_at__lte=previous_period_end, **filter_conditions)

    if len(previous_period_orders)==0:
        if not filter_conditions:
            period_orders_info = {
                'last_period_order_count': 0,
                'last_period_turnover': 0,
                'period_before_order_count': 0,
                'period_before_turnover': 0,
                'order_count_evolution_percentage': 0,
                'turnover_evolution_percentage': 0,
            }

        else:
            if filter_conditions == {'surplace': True}:
                period_orders_info = {
                    'last_period_order_count': 0,
                    'last_period_turnover': 0,
                    'period_before_order_count': 0,
                    'period_before_turnover': 0,
                    'order_count_evolution_percentage': 0,
                    'turnover_evolution_percentage': 0,
                    'last_period_on_site_out_of_total': (0 / get_periodicaly_total_orders()[
                        'previous_period_order_count']) * 100 if get_periodicaly_total_orders()[
                                                                     'previous_period_order_count'] != 0 else 0,
                    'period_before_on_site_out_of_total': (0 / get_periodicaly_total_orders()[
                        'previous_period_order_count']) * 100 if get_periodicaly_total_orders()[
                                                                     'previous_period_order_count'] != 0 else 0,
                }
            else:

                period_orders_info = {
                    'last_period_order_count': 0,
                    'last_period_turnover': 0,
                    'period_before_order_count': 0,
                    'period_before_turnover': 0,
                    'order_count_evolution_percentage': 0,
                    'turnover_evolution_percentage': 0,
                    'last_period_delivery_out_of_total': (0 / get_periodicaly_total_orders()[
                        'previous_period_order_count']) * 100 if get_periodicaly_total_orders()[
                                                                     'previous_period_order_count'] != 0 else 0,
                    'period_before_delivery_out_of_total': (0 / get_periodicaly_total_orders()[
                        'previous_period_order_count']) * 100 if get_periodicaly_total_orders()[
                                                                     'previous_period_order_count'] != 0 else 0,
                }
        return period_orders_info

    previous_period_order_count = previous_period_orders.count()
    previous_period_turnover = sum(order.total_price for order in previous_period_orders) or 0

    # Commandes de la periode d'avant
    period_before_orders = orders.objects.filter(
        create_at__gte=before_period_start, create_at__lte=before_period_end, **filter_conditions
    )
    period_before_order_count = period_before_orders.count()
    period_before_turnover = sum(order.total_price for order in period_before_orders) or 0

    # Calcul des évolutions en pourcentage
    order_count_evolution = 0
    turnover_evolution = 0

    if period_before_order_count > 0:
        order_count_evolution = ((
                                             previous_period_order_count - period_before_order_count) / period_before_order_count) * 100

    if period_before_turnover > 0:
        turnover_evolution = ((previous_period_turnover - period_before_turnover) / period_before_turnover) * 100

    # Retour des informations
    if not filter_conditions:

        period_orders_info = {
            'last_period_order_count': previous_period_order_count,
            'last_period_turnover': previous_period_turnover,
            'period_before_order_count': period_before_order_count,
            'period_before_turnover': period_before_turnover,
            'order_count_evolution_percentage': order_count_evolution,
            'turnover_evolution_percentage': turnover_evolution,
        }

    else:
        if filter_conditions == {'surplace': True}:
            period_orders_info = {
                'last_period_order_count': previous_period_order_count,
                'last_period_turnover': previous_period_turnover,
                'period_before_order_count': period_before_order_count,
                'period_before_turnover': period_before_turnover,
                'order_count_evolution_percentage': order_count_evolution,
                'turnover_evolution_percentage': turnover_evolution,
                'last_period_on_site_out_of_total': (previous_period_order_count / get_periodicaly_total_orders()[
                    'previous_period_order_count']) * 100 if get_periodicaly_total_orders()[
                                                                 'previous_period_order_count'] != 0 else 0,
                'period_before_on_site_out_of_total': (period_before_order_count / get_periodicaly_total_orders()[
                    'previous_period_order_count']) * 100 if get_periodicaly_total_orders()[
                                                                 'previous_period_order_count'] != 0 else 0,
            }
        else:

            period_orders_info = {
                'last_period_order_count': previous_period_order_count,
                'last_period_turnover': previous_period_turnover,
                'period_before_order_count': period_before_order_count,
                'period_before_turnover': period_before_turnover,
                'order_count_evolution_percentage': order_count_evolution,
                'turnover_evolution_percentage': turnover_evolution,
                'last_period_delivery_out_of_total': (previous_period_order_count / get_periodicaly_total_orders()[
                    'previous_period_order_count']) * 100 if get_periodicaly_total_orders()[
                                                                 'previous_period_order_count'] != 0 else 0,
                'period_before_delivery_out_of_total': (period_before_order_count / get_periodicaly_total_orders()[
                    'previous_period_order_count']) * 100 if get_periodicaly_total_orders()[
                                                                 'previous_period_order_count'] != 0 else 0,
            }

    return period_orders_info


def get_periodicaly_orders_by_type(period):
    # Commandes sur place
    on_site_info = get_periodicaly_orders_info(filter_conditions={'surplace': True}, period=period)

    # Commandes livrées
    delivery_info = get_periodicaly_orders_info(filter_conditions={'surplace': False}, period=period)

    return {
        'on_site_orders_info': on_site_info,
        'delivery_orders_info': delivery_info
    }


def get_periodicaly_delivery_infos():
    # Nombre total de livraisons effectuées
    total_delivery = {'last_period_order_count': get_periodicaly_orders_info(filter_conditions={'surplace': False})[
        'last_period_order_count'],  # la semaine passée
                      'period_before_order_count': get_periodicaly_orders_info(filter_conditions={'surplace': False})[
                          'period_before_order_count']}  # la semaine d'avant

    today = date.today()
    delivery_men_infos = []

    # Calcul des périodes
    previous_week_start = today - timedelta(days=today.weekday() + 7)
    previous_week_end = previous_week_start + timedelta(days=6)

    week_before_start = previous_week_start - timedelta(days=7)
    week_before_end = week_before_start + timedelta(days=6)

    delivery_men = DeliveryPerson.objects.all()

    if delivery_men:
        for deliMan in delivery_men:  # pour chaque livreur enregistré
            # Commandes de la semaine dernière
            previous_week_orders = orders.objects.filter(
                create_at__gte=previous_week_start, create_at__lte=previous_week_end,
                deliveryPerson_id=deliMan.id_deliveryman, status='delivered'
            )

            # Commandes de la semaine d'avant
            week_before_orders = orders.objects.filter(
                create_at__gte=week_before_start, create_at__lte=week_before_end,
                deliveryPerson_id=deliMan.id_deliveryman, status='delivered'
            )
            numb_delivery1 = len(previous_week_orders)
            total_pizzas1 = sum([order.pizza_and_extratopping_price for order in previous_week_orders if
                                 order.payment_method_order == 'delivered_man'])
            total_delivery1 = previous_week_orders.aggregate(
                total_delivery=Sum('deliveryPrice',
                                   filter=Q(deliveryPrice__isnull=False) & Q(payment_method_delivery='delivered_man'))
            )['total_delivery'] or 0

            numb_delivery2 = len(week_before_orders)
            total_pizzas2 = sum([order.pizza_and_extratopping_price for order in week_before_orders if
                                 order.payment_method_order == 'delivered_man'])
            total_delivery2 = week_before_orders.aggregate(total_delivery=Sum('deliveryPrice',
                                                                              filter=Q(deliveryPrice__isnull=False) & Q(
                                                                                  payment_method_delivery='delivered_man'))
                                                           )['total_delivery'] or 0

            delivDict = [deliMan.name, total_pizzas1, total_delivery1, 0.2 * total_delivery1, numb_delivery1,
                         total_pizzas2, total_delivery2, 0.2 * total_delivery2, numb_delivery2]
            delivery_men_infos.append(delivDict)

    return delivery_men_infos


def get_most_and_least_sold_pizza_names(period) -> dict:
    today = date.today()

    if period == "week":
        # Calcul des périodes
        previous_period_start = today - timedelta(days=today.weekday() + 7)
        previous_period_end = previous_period_start + timedelta(days=6)

        before_period_start = previous_period_start - timedelta(days=7)
        before_period_end = before_period_start + timedelta(days=6)

    else:
        month1 = 12 if today.month - 1 == 0 else today.month - 1  # le mois précedent 1,2,...,12
        month2 = month1 - 1  # le mois d'avant 1,2,...,12
        year = today.year - 1 if today.month - 1 == 0 else today.year  # l'année actuelle

        # Obtenir le premier jour du mois précédent
        previous_period_start = datetime(year, month1, 1)

        # Obtenir le dernier jour du mois précédent
        previous_period_end = datetime(year, month1,
                                       calendar.monthrange(year, month1)[1])

        # Obtenir le premier jour du mois d'avant
        before_period_start = datetime(year, month2, 1)

        # Obtenir le dernier jour du mois d'avant
        before_period_end = datetime(year, month2,calendar.monthrange(year, month2)[1])


    # Ventes de pizzas de la periode dernière
    previous_period_pizzas = Pizza.objects.filter(create_at__gte=previous_period_start, create_at__lte=previous_period_end)

    if len(previous_period_pizzas)==0:
        previous_period_list = None
    else:
        previous_period_sales = {}
        for pizza in previous_period_pizzas:
            pizza_name = pizza.get_name  # Utilisation de la propriété
            previous_period_sales[pizza_name] = previous_period_sales.get(pizza_name, 0) + 1

        previous_period_list = {name: count for name, count in previous_period_sales.items()}

    # Ventes de pizzas de la periode d'avant
    period_before_pizzas = Pizza.objects.filter(create_at__gte=before_period_start, create_at__lte=before_period_end)

    if len(period_before_pizzas) == 0:
        period_before_list = None
    else:
        period_before_sales = {}
        for pizza in period_before_pizzas:
            pizza_name = pizza.get_name  # Utilisation de la propriété
            period_before_sales[pizza_name] = period_before_sales.get(pizza_name, 0) + 1

        period_before_list = {name: count for name, count in period_before_sales.items()}

    # Retourner les résultats
    return {
        'previous_period': previous_period_list,
        'before_period': period_before_list
    }


"""def create_pdf_with_data(): # note: je dois ajouter un tableau en dessous qui resumme aussi les infos des livreurs
    second_table_data = get_periodicaly_delivery_infos()
    file_name = 'file.pdf'
    data = prepare_datas_to_export()
    if data:
        # Construire le chemin absolu du fichier PDF dans MEDIA_ROOT
        parent_path = os.path.join(settings.MEDIA_ROOT, 'reports')
        file_path = os.path.join(settings.MEDIA_ROOT, 'reports',file_name)

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=10)
        pdf.add_page()
        pdf.set_font("Arial", size=10)

        # Titre du document pour le premier tableau
        pdf.set_font("Arial", size=14, style='B')
        pdf.cell(0, 10, "Recapitulatif des commandes", ln=True, align='C')
        pdf.ln(10)  # Ajouter un espacement vertical

        # Entêtes du tableau
        headers = [
            "Date", "Sur place", "Adresse Livraison", "Heure livraison", "Livreur",
            "Statut", "Clients", "Mode de paiement commandes", "Mode de paiement livraison", "Infos Pizzas",
            "Prix Livraison", "Prix Pizzas + Suppléments", "Total"
        ]
        pdf.set_font("Arial", size=10, style='B')
        col_widths = [25, 20, 50, 25, 30, 25, 30, 35, 35, 40, 30, 45, 30]  # Largeurs des colonnes adaptées
        for col_num, header in enumerate(headers):
            pdf.cell(col_widths[col_num], 10, header, border=1, align='C')
        pdf.ln()  # Passer à la ligne suivante

        pdf.set_font("Arial", size=10, style='B')
        col_widths = [25, 20, 50, 25, 30, 25, 30, 35, 35, 40, 30, 45, 30]  # Largeurs des colonnes adaptées
        for col_num, header in enumerate(headers):
            pdf.cell(col_widths[col_num], 10, header, border=1, align='C')
        pdf.ln()  # Passer à la ligne suivante

        # Ajouter une nouvelle page pour le second tableau
        pdf.add_page()

        # Entêtes du second tableau
        second_headers = [
            "Livreurs", "Total vendu sem. passée", "Total livré sem. passée", "Pourcentage sem. passée",
            "Nombre livraisons sem. passée", "Total vendu sem. d'avant", "Total livré sem. d'avant",
            "Pourcentage sem. d'avant", "Nombre livraisons sem. d'avant"
        ]

        # Titre du document pour le second tableau
        pdf.set_font("Arial", size=14, style='B')
        pdf.cell(0, 10, "Statistiques des livreurs", ln=True, align='C')
        pdf.ln(10)  # Ajouter un espacement vertical

        col_widths_second = [30, 40, 40, 40, 45, 40, 40, 40, 45]  # Largeurs des colonnes adaptées
        pdf.set_font("Arial", size=10, style='B')
        for col_num, header in enumerate(second_headers):
            pdf.cell(col_widths_second[col_num], 10, header, border=1, align='C')
        pdf.ln()  # Passer à la ligne suivante

        pdf.set_font("Arial", size=10)
        for row_data in second_table_data:
            for col_num, value in enumerate(row_data):
                pdf.cell(col_widths_second[col_num], 10, str(value), border=1, align='C')
            pdf.ln()  # Passer à la ligne suivante

        # Sauvegarder le fichier PDF
        pdf.output(file_path)

        return  parent_path, file_path"""


def plot_empty_polar(file_name, fig,ax):
    # Listes vides
    categories = []
    values = []

    # Création de l'axe polaire
    fig, ax = plt.subplots(subplot_kw={'projection': 'polar'})

    # Tentative de traçage
    try:
        angles = np.linspace(0, 2 * np.pi, len(categories), endpoint=False).tolist()
        values += values[:1]  # Répéter la première valeur pour fermer le graphique
        angles += angles[:1]
        ax.fill(angles, values, color='blue', alpha=0.25)
        ax.plot(angles, values, color='blue', linewidth=2)
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(categories)
        ax.set_title('Aucune vente enregistrée', fontsize=14)


    except Exception as e:
        print(f"Erreur : {e}")

    finally:
        plt.savefig(os.path.join(img_path, file_name), dpi=300, bbox_inches="tight")
        return os.path.join(img_path, file_name)



def generate_2x_charts(datalist, filename='Total pizzas vendues.pdf'):

    if datalist[0]:
        data1 = datalist[0][0]
    else:
        data1 = pd.DataFrame()

    if datalist[1]:
        data2 = datalist[1][0]
    else:
        data2 = pd.DataFrame()


    # Créer une figure avec deux sous-graphiques polaires côte à côte
    fig, axes = plt.subplots(2, 1, figsize=(12, 18))

    if data1.empty and data2.empty:
        """categories = []
        values = []
        empty = pd.DataFrame({'Pizzas':categories, 'Ventes':values})
        try:

            ax1 = axes[0]
            sns.barplot(legend=False, ax=ax1, x='Pizzas', y='Ventes', data=empty, ci=None)
            ax1.title(datalist[0][1])
            ax1.xlabel("Pizzas")
            ax1.ylabel("Ventes")


            ax2 = axes[1]
            sns.barplot(legend=False, ax=ax2, x='Pizzas', y='Ventes', data=empty, ci=None)
            ax2.title(datalist[1][1])
            ax2.xlabel("Pizzas")
            ax2.ylabel("Ventes")


        except Exception as e:
            print(f"Erreur : {e}")

        finally:
            # Ajuster l'espacement
            plt.tight_layout()

            # Sauvegarde dans un fichier PDF
            with PdfPages(os.path.join(img_path, filename)) as pdf:
                pdf.savefig(fig)  # Sauvegarde la figure
                plt.close(fig)

            return os.path.join(img_path, filename)"""

        return None

    elif data1.empty:

        norm2 = plt.Normalize(data2['Ventes'].min(), data2['Ventes'].max())
        colors2 = plt.cm.coolwarm(norm2(data2['Ventes']))
        sns.barplot(legend=False, x='Ventes', y='Pizzas', data=data2, errorbar=None, palette=colors2.tolist(), orient='h')

        # Ajouter des titres et des étiquettes
        plt.title(datalist[1][1])
        plt.ylabel("Pizzas")
        plt.xlabel("Ventes")

        # Enregistrer le graphique dans un fichier PDF
        plt.savefig(os.path.join(img_path, filename), format="pdf")

        return os.path.join(img_path, filename)

    elif data2.empty:

        norm1 = plt.Normalize(data1['Ventes'].min(), data1['Ventes'].max())
        colors1 = plt.cm.coolwarm(norm1(data1['Ventes']))
        sns.barplot(legend=False, x='Ventes', y='Pizzas', data=data1, errorbar=None, palette=colors1.tolist(), orient='h')

        # Ajouter des titres et des étiquettes
        plt.title(datalist[0][1])
        plt.ylabel("Pizzas")
        plt.xlabel("Ventes")

        # Enregistrer le graphique dans un fichier PDF
        plt.savefig(os.path.join(img_path, filename), format="pdf")

        return os.path.join(img_path, filename)

    else:
        norm1 = plt.Normalize(data1['Ventes'].min(), data1['Ventes'].max())
        colors1 = plt.cm.coolwarm(norm1(data1['Ventes']))

        ax1 = axes[0]
        sns.barplot(legend=False, ax=ax1, x='Ventes', y='Pizzas', data=data1, errorbar=None, palette=colors1.tolist(), orient='h')
        ax1.title(datalist[0][1])
        ax1.ylabel("Pizzas")
        ax1.xlabel("Ventes")

        norm2 = plt.Normalize(data2['Ventes'].min(), data2['Ventes'].max())
        colors2 = plt.cm.coolwarm(norm2(data2['Ventes']))

        # Deuxième graphique polaire
        ax2 = axes[1]
        sns.barplot(legend=False, ax=ax2, x='Ventes', y='Pizzas', data=data2, errorbar=None, palette=colors2.tolist(), orient='h')
        ax2.title(datalist[1][1])
        ax1.ylabel("Pizzas")
        ax1.xlabel("Ventes")


        # Ajuster l'espacement
        plt.tight_layout()

        # Sauvegarde dans un fichier PDF
        with PdfPages(os.path.join(img_path, filename)) as pdf:
            pdf.savefig(fig)  # Sauvegarde la figure
            plt.close(fig)

        return os.path.join(img_path,filename)


def generate_4x_charts(datasets, file_name="total commandes et chiffres d'affaire.pdf"):

    # Configuration de la grille
    fig, axes = plt.subplots(2, 2, figsize=(18, 18))
    axes = axes.flatten()

    for i, data in enumerate(datasets):
        df = pd.DataFrame(data[0])

        # Position des barres pour chaque jeu de données
        x = np.arange(len(df))  # Position des catégories
        width = 0.35  # Largeur des barres

        # Barplot avec deux jeux de données côte à côte
        axes[i].bar(x - width / 2, df["dataset1"], width, label=data[3], color="cornflowerblue")
        axes[i].bar(x + width / 2, df["dataset2"], width, label=data[4], color="tomato")

        axes[i].set_title(data[5], fontsize=14)
        if data[1]:
            axes[i].set_xlabel(data[1], fontsize=12)
        axes[i].set_ylabel(data[2], fontsize=12)
        axes[i].set_xticks(x)
        axes[i].set_xticklabels(df["Catégorie"], fontsize=12)
        axes[i].legend()

        axes[i].set_ylim(0, max(df["dataset1"].max(), df["dataset2"].max()) + 5)

        # Ajouter des annotations de valeur sur les barres
        for j, val in enumerate(df["dataset1"]):
            axes[i].text(x[j] - width / 2, val + 0.5, f"{val}", ha="center", fontsize=10)
        for j, val in enumerate(df["dataset2"]):
            axes[i].text(x[j] + width / 2, val + 0.5, f"{val}", ha="center", fontsize=10)

    # Ajuster l'espacement
    plt.tight_layout()

    # Sauvegarde dans un fichier PDF
    with PdfPages(os.path.join(img_path,file_name)) as pdf:
        pdf.savefig(fig)

    return os.path.join(img_path,file_name)


def create_excel_with_data(file_name,first_period=None,second_period=None):
    data = prepare_datas_to_export(first_period,second_period)
    if data:
        # Construire le chemin absolu du fichier Excel dans MEDIA_ROOT
        file_path = os.path.join(settings.MEDIA_ROOT, 'reports',file_name)

        workbook = openpyxl.Workbook()
        # Obtenir la feuille active par défaut
        sheet = workbook.active

        sheet.title = "Recapitulatif des commandes"

        headers = [
            "Date", "Sur place", "Adresse Livraison", "Heure livraison", "Livreur",
            "Statut", "Clients", "Mode de paiement commandes", "Mode de paiement livraison", "Infos Pizzas",
            "Prix Livraison", "Prix Pizzas + Suppléments", "Total"
        ]

        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)

        # Insérer les données (à partir de la ligne 2)
        start_row = 2
        for row_num, row_data in enumerate(data, start=start_row):
            for col_num, value in enumerate(row_data, start=1):
                sheet.cell(row=row_num, column=col_num, value=value)

        # Sauvegarder le fichier Excel
        workbook.save(file_path)

        return file_path
