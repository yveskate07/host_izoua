import json
from django.middleware.csrf import get_token
from bs4 import BeautifulSoup
from django.contrib.auth.views import LoginView, LogoutView
from django.urls import reverse
from django.utils.timezone import now
from django.shortcuts import render, redirect
from django.conf import settings
from accounts.auth_form import UserLoginForm
from .models import orders, Pizza, Client, DailyInventory, ExtraTopping, PizzaSizePrice, DeliveryPerson, PizzaName
from datetime import timedelta, datetime, date
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import os
import locale
from django.db.models import Sum
from django.db.models import Sum, F, Q

locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')


# Create your views here.
@login_required
def add_inventory(request): # fonction qui cree et ajoute un nouvel inventaire dans la bd
    if request.method == "POST":
        date = datetime.strptime(request.POST.get('addDate'),"%Y-%m-%d").date()
        grande = int(request.POST.get('addGrande'))
        mini = int(request.POST.get('addMini'))

        DailyInventory.objects.create(small_pizzas_count=mini,large_pizzas_count=grande,date=date)

        return redirect(reverse('home'))

    return redirect(reverse('home'))

class IzouaLoginView(LoginView):
    form_class = UserLoginForm

class IzouaLogoutView(LogoutView):
    next_page = '/login'

@login_required
def to_admin(request):
    return redirect('/admin/')

@login_required
def filter_orders_by_status(request): # quand l'utilisateur change de filtre

    if request.method == 'POST':
        return fetching_datas(request, filter_=request.POST.get('selected_option'), date_to_print=request.session['date_selected'])

    return redirect(reverse('home'))


@login_required
def filter_orders_by_date(request): # quand l'utilisateur change de date à partir du calendrier date_from_form

    if request.method == 'POST':
        request.session['date_selected'] = request.POST.get('datePicker')
        return fetching_datas(request, filter_=date.fromisoformat(request.session['date_selected']),
                       date_to_print=datetime.strptime(request.POST.get('datePicker'), "%Y-%m-%d").date().isoformat())

    return redirect(reverse('home'))


def split_html_and_get_pizzas(html, pizzas_count_sold, firstOrderStatus, finalOrderStatus, edit=False):

    # html: un gros bloc de code html contenant toute les pizzas,
    # pizzas_count_sold: dictionnaire contenant le nombre de petites/grandes pizzas vendues

    order_liste = html.split('///') # liste de code html où chaque code est un ensemble d'inputs contenant les infos sur une pizzas achetée: nom, taille, suppléments, prix

    pizza_liste = []

    for code_html in order_liste: # code_html: code html d'une pizza
        soup = BeautifulSoup(code_html, 'html.parser')
        inputs = soup.find_all("input") # liste d'inputs nom, taille, suppléments, prix pour une pizza choisie

        pizza_dict = dict()
        status = ''

        if len(inputs) > 0:
            for input_tag in inputs: # input_tag ici represente un input qui contient soit nom soit taille soit suppléments soit prix
                if input_tag.get("id") == 'name': # si input_tag est un input contenant la valeur nom
                    if " - " in input_tag.get("value"):
                        pizza_dict['name'] = input_tag.get("value").split(" - ")  # name = ['moitie 1', 'moitie 2']
                        status = 'speciale'
                    else:
                        pizza_dict['name'] = input_tag.get("value")
                        status = 'normale'
                if input_tag.get("id") == 'size': # si input_tag est un input contenant la valeur taille
                    pizza_dict['size'] = input_tag.get("value")
                    if not edit: # si on n'est pas en mode edition, on est donc en train d'ajouter une nouvelle pizza, donc le solde de pizza vendue de cette taille doit etre incrémenté
                        pizzas_count_sold[input_tag.get("value")] += 1

                if input_tag.get("id") == 'extratoppings': # si input_tag est un input contenant la valeur suppléments
                    if len(input_tag.get("value").split(', ')) > 1:
                        pizza_dict['extratoppings'] = input_tag.get("value").split(', ')
                    else:
                        pizza_dict['extratoppings'] = []
                if input_tag.get("id") == 'price': # si input_tag est un input contenant la valeur prix
                    pizza_dict['price'] = int(input_tag.get("value"))

            if status == 'speciale':
                extratoppings = (ExtraTopping.objects.filter(name=i)[0] for i in pizza_dict['extratoppings'])
                pizza = Pizza.objects.create(
                    moitie_1=pizza_dict['name'][0],
                    moitie_2=pizza_dict['name'][1],
                    status='Spéciale',
                    name='Pizza Spéciale',
                    size=pizza_dict['size'])
                pizza.extratoppings.add(*extratoppings)
                pizza_liste.append(pizza)
            else:
                if len(pizza_dict['extratoppings']) > 0:
                    extratoppings = tuple(ExtraTopping.objects.filter(name=i)[0] for i in pizza_dict['extratoppings'])
                pizza = Pizza.objects.create(
                    moitie_1='Néant',
                    moitie_2='Néant',
                    status='Normale',
                    name=pizza_dict['name'],
                    size=pizza_dict['size'])
                if len(pizza_dict['extratoppings']) > 0:
                    pizza.extratoppings.add(*extratoppings)
                pizza_liste.append(pizza)

    if not edit: # si on n'est pas en mode edition
        return pizza_liste, pizzas_count_sold

    return pizza_liste


def get_pizzas_names_from_html_input(html):
    order_liste = html.split('///')
    pizza_liste_names = []

    for code_html in order_liste:
        soup = BeautifulSoup(code_html, 'html.parser')
        inputs = soup.find_all("input")

        if len(inputs) > 0:
            for input_tag in inputs:
                if input_tag.get("id") == 'name':
                    pizza_liste_names.append(input_tag.get('value'))

    return ", ".join(pizza_liste_names)

@login_required
def edit_order_if_granted(request):
    if request.method == 'POST':
        order_to_edit = int(request.POST.get('grantedEdit'))
        with open('izouapp/static/izouapp/data.json', 'r') as file:
            content = json.load(file)
            for dict_ in content['pending_to_edit']:
                if dict_['order_id'] == order_to_edit:
                    data = dict_['data']
                    break
            edit_order_directly(request,
                                client_to_edit=data[1], # int
                                order_to_edit=data[0], # int
                                id_deliveryman=data[2], # int
                                deliveryman_name=data[3], # str
                                order_type=data[4], # str
                                html_list_order=data[7], # str
                                client_name=data[5], # str
                                client_number=data[8], # str
                                client_adress=data[9], # str
                                delivery_time=data[10], # str
                                payment_method_from_order_to_deliver=data[11], # str
                                order_status=data[6], # str
                                price_delivery=data[12],
                                pizzas_count_av = data[13]# int
                                )
    return redirect(reverse('home'))

def edit_order_directly(request,client_to_edit,order_to_edit,id_deliveryman,deliveryman_name,order_type,html_list_order,client_name,client_number,client_adress,delivery_time,payment_method_from_order_to_deliver,order_status,price_delivery, pizzas_count_av):

    client = Client.objects.filter(id_client=client_to_edit)
    order = orders.objects.filter(order_id=order_to_edit)

    obj_first = order.first()

    initial_status = obj_first.status

    current_inventory = DailyInventory.objects.filter(date=get_date(request))

    user = request.user

    if user.is_admin:

        for pizza in obj_first.pizzas.all():
            pizza.delete()  # supprime toutes les pizzas liées à cette commande

        if order_type == 'to-deliver':

            if DeliveryPerson.objects.filter(id_deliveryman=id_deliveryman).first() != DeliveryPerson.objects.filter(
                    name=deliveryman_name).first():  # si l'id du livreur et le nom du livreur livreur ne correspondent pas#

                deliv_man = DeliveryPerson.objects.filter(name=deliveryman_name).first()

            else:
                deliv_man = DeliveryPerson.objects.filter(id_deliveryman=id_deliveryman).first()

            client.update(name=client_name, phone_number=client_number, adress=client_adress)

            obj_first.status = order_status
            obj_first.deliveryHour = delivery_time
            obj_first.deliveryAdress = client_adress
            obj_first.payment_method = payment_method_from_order_to_deliver
            obj_first.update_at = now()
            obj_first.surplace = False
            obj_first.deliveryPerson = deliv_man
            obj_first.client = client.first()
            obj_first.deliveryPrice = price_delivery

            pizzas = split_html_and_get_pizzas(html_list_order, pizzas_count_sold=None, edit=True, finalOrderStatus=order_status, firstOrderStatus=initial_status)

        else:

            client.update(name=client_name)
            obj_first.update_at = now()
            obj_first.surplace = True
            obj_first.status = order_status if order_status else 'on-site'
            obj_first.client = client.first()

            pizzas = split_html_and_get_pizzas(html_list_order, pizzas_count_sold=None, edit=True, finalOrderStatus=order_status, firstOrderStatus=initial_status)

        obj_first.pizzas.add(*tuple(pizzas))
        obj_first.save()
        order.update(edit_requested=False)
        current_inventory.update(sold_small_pizzas_count=current_inventory[0].small_pizzas_count-pizzas_count_av['Petite'],
                                 sold_large_pizzas_count=current_inventory[0].large_pizzas_count-pizzas_count_av['Grande'])  # mise à jour de l'inventaire


@login_required
def edit_order(request): # modifie simplement le dictionnaire data.json et la clé pending_to_edit
    if request.method == 'POST':

        order_to_edit = int(request.POST.get('order-id'))
        client_to_edit = int(request.POST.get('client-id'))

        order_type = request.POST.get('order-type')

        client_name = request.POST.get('client_name')

        order_status = request.POST.get('editOrderStatus')

        if order_type == 'to-deliver':

            html_list_order = request.POST.get('edit-hidden-textarea-from-order-on-delivery1')
            client_number = request.POST.get('client_number')
            client_adress = request.POST.get('client_adress')
            id_deliveryman = int(request.POST.get('deliver-id'))

            deliveryman_name = request.POST.get('delivery_man')
            delivery_time = request.POST.get('delivery_time')

            pizzas_count_av = {'Petite': int(request.POST.get('edit-SmallPizzasAvailableOnDelivery')), # les pizzas encore disponibles
                               'Grande': int(request.POST.get('edit-LargePizzasAvailableOnDelivery'))} # les pizzas encore disponibles

            if DeliveryPerson.objects.filter(id_deliveryman=id_deliveryman).first() != DeliveryPerson.objects.filter(
                    name=deliveryman_name).first():  # si l'id du livreur et le nom du livreur livreur ne correspondent pas#

                deliv_man = DeliveryPerson.objects.filter(name=deliveryman_name).first()

            else:
                deliv_man = DeliveryPerson.objects.filter(id_deliveryman=id_deliveryman).first()

            payment_method_from_order_to_deliver = request.POST.get('payment_method_from_order_to_deliver')
            price_delivery = int(request.POST.get('price_delivery'))
        else:
            html_list_order = request.POST.get('edit-hidden-textarea-from-order-on-site1')

            pizzas_count_av = {'Petite': int(request.POST.get('edit-SmallPizzasAvailableOnSite')), # les pizzas encore disponibles
                               'Grande': int(request.POST.get('edit-LargePizzasAvailableOnSite'))} # les pizzas encore disponibles

            deliveryman_name = 'Commandé sur place'
            client_number = ''
            client_adress = ''
            delivery_time = ''
            payment_method_from_order_to_deliver = ''
            id_deliveryman = None
            price_delivery = ''

        client = Client.objects.filter(id_client=client_to_edit)
        order = orders.objects.filter(order_id=order_to_edit)

        obj_first = order.first()

        initial_status = obj_first.status

        current_inventory = DailyInventory.objects.filter(date=get_date(request))

        user = request.user

        if user.is_admin:

            obj_first.pizzas.all().delete() # supprime toutes les pizzas liées à cette commande


            if order_type == 'to-deliver':

                pizzas = split_html_and_get_pizzas(html_list_order, pizzas_count_sold=None, edit=True, firstOrderStatus=initial_status, finalOrderStatus=order_status)


                client.update(name=client_name,phone_number=client_number,adress=client_adress)
                obj_first.status = order_status
                obj_first.deliveryHour = delivery_time
                obj_first.deliveryAdress = client_adress
                obj_first.payment_method = payment_method_from_order_to_deliver
                obj_first.update_at = now()
                obj_first.surplace = False
                obj_first.deliveryPerson = deliv_man
                obj_first.client = client.first()
                obj_first.deliveryPrice = price_delivery

            else:
                pizzas = split_html_and_get_pizzas(html_list_order, pizzas_count_sold=None, edit=True, firstOrderStatus=initial_status, finalOrderStatus=order_status)

                client.update(name=client_name)
                obj_first.update_at=now()
                obj_first.surplace=True
                obj_first.status=order_status
                obj_first.client=client.first()

            obj_first.pizzas.add(*tuple(pizzas))
            obj_first.save()

            current_inventory.update(sold_small_pizzas_count=current_inventory[0].small_pizzas_count-pizzas_count_av['Petite'],
                                     sold_large_pizzas_count=current_inventory[0].large_pizzas_count-pizzas_count_av['Grande'])  # mise à jour de l'inventaire

        else:

            pizza_names = get_pizzas_names_from_html_input(html_list_order)

            csrf_token = get_token(request)

            if order_type == 'to-deliver':
                pizzas_count_av = {'Petite': int(request.POST.get('edit-SmallPizzasAvailableOnDelivery')),
                                   'Grande': int(request.POST.get(
                                       'edit-LargePizzasAvailableOnDelivery'))}

            else:
                pizzas_count_av = {'Petite': int(request.POST.get('edit-SmallPizzasAvailableOnSite')),
                                   'Grande': int(request.POST.get(
                                       'edit-LargePizzasAvailableOnSite'))}

            with open('izouapp/static/izouapp/data.json', 'r') as file:
                content = json.load(file)
                len_row = len(content['pending_to_edit'])
                content['pending_to_edit'].append({'order_id':order_to_edit,'idRow':f'pending-{len_row}','trFirst':f"""
                                                   <td class="text-center">{len_row+1}</td>
                                                   <td class="text-center">{client[0].__str__()}</td>
                                                   <td class="text-center">{pizza_names}</td>
                                                   <td class="text-center">{'<span class="badge bg-success">Livrée</span>' if order_status=='delivered' else '<span class="badge bg-danger">Annulée</span>' if order_status=='canceled' else '<span class="badge bg-secondary">Sur place</span>'}</td>
                                                   <td class="text-center">{obj_first.total_price}</td>
                                                   <td class="text-center">{obj_first.create_at}</td>
                                                   <td class="text-center">{obj_first.deliveryPerson.name if obj_first.deliveryPerson else 'Commandé sur place'}</td>"""
                                                   ,'formInnerHTML':f"""
                                                   <input type="hidden" name="grantedEdit" value="{order_to_edit}" class="form-control">
                                                   <input type='submit' value='Valider' class='confirmPending form-control' disabled style="background-color: green; padding: 10px; border-radius: 5px; display: inline-block;">""",
                                                   'data':[order_to_edit, # int
                                                            client_to_edit, # int
                                                            id_deliveryman, # int
                                                            deliveryman_name, # str
                                                            order_type, # str
                                                            client_name, # str
                                                            order_status, # str
                                                            html_list_order, # str
                                                            client_number, # str
                                                            client_adress, # str
                                                            delivery_time, # str
                                                            payment_method_from_order_to_deliver, # str
                                                            price_delivery,
                                                            pizzas_count_av # dict
                                                           ]}) # int

            with open('izouapp/static/izouapp/data.json', 'w') as file:
                json.dump(content, file)

            order.update(edit_requested=True)

    return redirect(reverse('home'))


@login_required
def home(request): # quand l'utilisateur atterit sur la page pour la premiere fois

    datas_to_json(request)

    return fetching_datas(request, filter_=None, date_to_print=now().date().isoformat())


def fetching_datas(request, filter_, date_to_print):

    time_now = datetime.now().strftime("%H:%M:%S")

    context = dict()
    request.session['date_selected'] = date_to_print # un objet de type str et non date

    fetched_datas_clients = []
    fetched_datas_inventory = []

    list_order = []
    list_inventory = []
    list_client = []

    if isinstance(filter_, date): # dans le cas où le filtre est un objet de type date
        fetched_datas_orders = orders.objects.filter(
            create_at=filter_)

    elif isinstance(filter_, str): # dans le cas où le filtre est un objet de type str: all ou canceled ou delivered
        context['preselected_filter'] = {'all':'Tout', 'canceled':'Annulés','delivered':'Livrés'}.get(filter_) # retourne le contenu html du filtre selectionné

        if filter_ in ['canceled','delivered']:
            fetched_datas_orders = orders.objects.filter(create_at=date.fromisoformat(request.session['date_selected']),
                                                         status=filter_)

        else:
            fetched_datas_orders = orders.objects.filter(create_at=date.fromisoformat(request.session['date_selected']))

    else:
        context['preselected_filter'] = None
        fetched_datas_orders = orders.objects.filter(create_at=date.fromisoformat(request.session['date_selected'])) # récupère toutes les commandes créées à la date fournie en paramètre
    for i in range(7):
        daily_inventory = DailyInventory.objects.filter(
            date=date.fromisoformat(request.session['date_selected']) - timedelta(i))
        if daily_inventory:
            fetched_datas_inventory.append(daily_inventory[0]) # récupère l'historique de l'inventaire des 7 derniers jours
        else:
            continue

    if fetched_datas_orders: # si la liste des commandes du jour choisi n'est pas vide
        with open('izouapp/static/izouapp/data.json', 'r') as file:
            content = json.load(file)
            html_list_order = []
            for data in fetched_datas_orders: # pour chaque commande de ce jour
                client = Client.objects.filter(id_client=data.client.id_client).first() # le client auquel cette commande est associée
                if not data.edit_requested: # si la commande n'est pas en cours de modification
                    if client: # si le client existe
                        fetched_datas_clients.append(client) # liste des clients ayant commandé ce jour
                        pizza_names = ", ".join([" - ".join([pizza.moitie_1, pizza.moitie_2]) if pizza.status == 'Spéciale' else pizza.name for pizza in data.pizzas.all()])
                        #pizza_price = sum([pizza.price for pizza in data.pizzas.all()])
                        order = {'a': data.order_id,'b': client.name, 'c': pizza_names, 'd': data.status, 'e': data.total_price,
                                 'f': data.create_at,
                                 'g': data.deliveryPerson if data.surplace == False else 'Commandé sur place',
                                 'h': data.surplace, 'i':data.client.id_client, 'j':data.deliveryPerson.id_deliveryman  if data.surplace == False else None,
                                 'k':sum([1 for i in data.pizzas.all() if i.size=='Petite']), 'l':sum([1 for i in data.pizzas.all() if i.size=='Grande']),'m':client.adress, 'n':client.phone_number} # dictionnaire des commandes

                    else: # si le client n'existe pas
                        print("le client n'existe pas")
                        fetched_datas_clients.append(client)
                        pizza_names = ", ".join(
                            [" - ".join([pizza.moitie_1, pizza.moitie_2]) if pizza.status == 'Spéciale' else pizza.name
                             for pizza in data.pizzas.all()])
                        # pizza_price = sum([pizza.price for pizza in data.pizzas.all()])
                        order = {'a': data.order_id, 'b': '', 'c': pizza_names, 'd': data.status,
                                 'e': data.total_price,
                                 'f': data.create_at,
                                 'g': data.deliveryPerson if data.surplace == False else 'Commandé sur place',
                                 'h': data.surplace, 'i': -1, 'j':data.deliveryPerson.id_deliveryman  if data.surplace == False else None}
                    list_order.append(order)

                for pizza in data.pizzas.all(): # pour chaque pizza commandé dans la commande actuelle
                    if pizza.status == 'Spéciale':

                        # code html qui sera affiché pour la pizza actuelle
                        html_pizza = f"""<div class="row my-1"><input type="text" class='col-md-2 mx-1' id="name" value="{pizza.moitie_1+' - '+pizza.moitie_2}" disabled><input type="text" class='col-md-2 mx-1' id="size" value="{pizza.size}" disabled><input type="text" class='col-md-2 mx-1' id="extratoppings" value="{", ".join(
                            [extratopping.name for extratopping in pizza.extratoppings.all()]) if len(pizza.extratoppings.all())>0 else ''}" disabled><input type="text" class='col-md-2 mx-1' id="price" value="{pizza.price}" disabled><button class="btn btn-danger col-md-4" style="width: 20%; margin: 8px" onclick="deleteItemFromInvoice(taille='{pizza.size}',event, """+f"""'{'edit-summary-box-from-order-on-site' if data.surplace else 'edit-summary-box-from-order-on-delivery'}')" type="button">Supprimer</button></div>"""

                        html_list_order.append(html_pizza) # liste de code html des pizzas, donc constituant le code html de la commande

                    else:

                        html_pizza = f"""<div class="row my-1"><input type="text" class='col-md-2 mx-1' id="name" value="{pizza.name}" disabled><input type="text" class='col-md-2 mx-1' id="size" value="{pizza.size}" disabled><input type="text" class='col-md-2 mx-1' id="extratoppings" value="{", ".join(
                            [extratopping.name for extratopping in pizza.extratoppings.all()]) if len(pizza.extratoppings.all()) > 0 else ''}" disabled><input type="text" class='col-md-2 mx-1' id="price" value="{pizza.price}" disabled><button class="btn btn-danger col-md-4" style="width: 20%; margin: 8px" onclick="deleteItemFromInvoice(taille='{pizza.size}',event, """+f"""'{'edit-summary-box-from-order-on-site' if data.surplace else 'edit-summary-box-from-order-on-delivery'}')" type="button">Supprimer</button></div>"""

                        html_list_order.append(html_pizza)
                    # à la fin html_list_order contient une liste de code html pour les pizzas speciales et normales mais pour la commande en cours

                if 'orderToHtml' in content.keys():
                    for dict_ in  content['orderToHtml']: # à la recherche du dictionnaire correcpondant à la commande actuelle dans le fichier data.json
                        if dict_['order_id'] == data.order_id: # si trouvé
                            if not data.surplace:
                                content['orderToHtml'][content['orderToHtml'].index(dict_)] = {'order_id':data.order_id,'client_infos': {'name': client.name, 'number': client.phone_number,
                                                                                                                                         'deliveryAdress': data.deliveryAdress,
                                                                                                                                         'paymentMode': data.payment_method,
                                                                                                                                         'delivPrice': data.deliveryPrice,
                                                                                                                                         'deliveryMan': data.deliveryPerson.name,
                                                                                                                                         'deliveryHour': data.deliveryHour.strftime("%H:%M")},
                                                                                               'orderHTML': "".join(html_list_order)}
                            else:
                                content['orderToHtml'][content['orderToHtml'].index(dict_)] = {'order_id': data.order_id, 'client_infos': {'name': client.name},'orderHTML': "".join(html_list_order)}
                            break

                html_list_order = []

            with open('izouapp/static/izouapp/data.json', 'w') as file:
                json.dump(content, file)

        for client in set(fetched_datas_clients):
            list_client.append({0: client.id_client, 1: client.name, 2: client.phone_number, 3: client.adress})


        context['orders'] = list_order
        context['clients'] = list_client
        context['len_orders'] = len(context['orders'])
        context['len_orders_delivered'] = len([i for i in context['orders'] if i['d'] == 'delivered'])
        context['len_orders_canceled'] = len([i for i in context['orders'] if i['d'] == 'canceled'])

    else:
        context['orders'] = []
        context['clients'] = []
        context['len_orders'] = 0
        context['len_orders_delivered'] = 0
        context['len_orders_canceled'] = 0

    if fetched_datas_inventory:
        for data in fetched_datas_inventory:
            inventory = {1: data.date, 2: data.large_pizzas_count-data.sold_large_pizzas_count,
                         3: data.small_pizzas_count-data.sold_small_pizzas_count, 4: data.sold, 5: data.remaining}
            list_inventory.append(inventory)
        context['inventories'] = list_inventory
    else:
        context['inventories'] = []
        
    delivery_men = DeliveryPerson.objects.all()
    context['delivery_men_infos'] = []
    
    if delivery_men:
        for deliMan in delivery_men:
            orders_ = orders.objects.filter(
                        deliveryPerson_id=deliMan.id_deliveryman,
                        create_at=request.session['date_selected']
                    )
            total_pizzas = sum([order.pizza_and_extratopping_price for order in orders_])
            total_delivery = orders_.aggregate(
            total_delivery=Sum('deliveryPrice', filter=Q(deliveryPrice__isnull=False))
        )['total_delivery'] or 0
            print('total_delivery: ', total_delivery)
            
            delivDict = {'name':deliMan.name, 'TotalPizzasSold':total_pizzas, 'TotalDeliv':total_delivery, 'Percent':0.2*total_delivery}
            context['delivery_men_infos'].append(delivDict)

    preselected_datetime = datetime.strptime(
        f"{request.session.get('date_selected')} {time_now}", "%Y-%m-%d %H:%M:%S").isoformat()
    context['preselected_datetime'] = preselected_datetime # la date qui sera affichée dans le calendrier
    context['user'] = request.user


    return render(request, 'izouapp/manager_screen.html', context=context)


@login_required
def add_order(request):

    if request.method == 'POST':

        with open('izouapp/static/izouapp/data.json', 'r') as file:
            content = json.load(file)

            current_inventory = DailyInventory.objects.filter(date=get_date(request)) # inventaire de la date actuelle

            pizzas_count_sold = {'Petite': current_inventory[0].sold_small_pizzas_count,  # le nombre de petites pizzas vendues
                                  'Grande': current_inventory[0].sold_large_pizzas_count} # le nombre de grandes pizzas vendues

            if request.POST.get('addOrder') == 'order_to_deliver': # si la commande est à livrer
                html_list_order = request.POST.get('hidden-textarea-from-order-on-delivery') # bloc html de toutes les pizzas + le bouton supprimer à reintroduire dans le html

                infos_client = {'name': request.POST.get('client_name'),
                                'numero':request.POST.get('client_number'),
                                'adresse': request.POST.get('client_adress'),
                                'methode_payement': request.POST.get('payment_method_from_order_to_deliver'),
                                'prix_livraison': request.POST.get('price_delivery'),
                                'livreur': request.POST.get('delivery_man'),
                                'heure_livraison': request.POST.get('delivery_time')}

                order_html = request.POST.get('hidden-textarea-from-order-on-delivery1') # bloc html de toutes les pizzas

                pizzas, pizzas_count_sold = split_html_and_get_pizzas(html=order_html, pizzas_count_sold=pizzas_count_sold, firstOrderStatus='delivered', finalOrderStatus='delivered') # noms des pizzas commandées + inventaire des pizzas vendues mis à jour
                client = Client.objects.create( # création d'un nouveau client
                    name=infos_client['name'],
                    phone_number=infos_client['numero'],
                    adress=infos_client['adresse'])


                order = orders.objects.create( # création d'une nouvelle commande
                    deliveryHour = infos_client['heure_livraison'],
                    deliveryAdress = infos_client['adresse'],
                    payment_method = infos_client['methode_payement'],
                    deliveryPerson = DeliveryPerson.objects.filter(name = infos_client['livreur'])[0],
                    client = client,
                    deliveryPrice = infos_client['prix_livraison'],
                )

                order.pizzas.add(*tuple(pizzas)) # ajout des pizzas dans la commande

                current_inventory.update(sold_small_pizzas_count=pizzas_count_sold['Petite'],
                                         sold_large_pizzas_count=pizzas_count_sold['Grande']) # mise à jour de l'inventaire

                content['orderToHtml'].append({'order_id':order.order_id,'client_infos':{'name':infos_client['name'],'number':infos_client['numero'],
                                              'deliveryAdress':infos_client['adresse'],'paymentMode':infos_client['methode_payement'],
                                              'delivPrice':infos_client['prix_livraison'],'deliveryMan':infos_client['livreur'],
                                              'deliveryHour':infos_client['heure_livraison']},
                                              'orderHTML':html_list_order})

            else: # si la commande est sur place

                html_list_order = request.POST.get('hidden-textarea-from-order-on-site')

                infos_client = {'name_onsite':request.POST.get('client_name')}
                order_html = request.POST.get('hidden-textarea-from-order-on-site1')

                pizzas, pizzas_count_sold = split_html_and_get_pizzas(html=order_html,
                                                                      pizzas_count_sold=pizzas_count_sold, firstOrderStatus='delivered', finalOrderStatus='delivered')

                client = Client.objects.create(
                    name=infos_client['name_onsite'])

                order = orders.objects.create(
                    surplace = True,
                    client=client,
                )

                order.pizzas.add(*tuple(pizzas))

                current_inventory.update(sold_small_pizzas_count=pizzas_count_sold['Petite'],sold_large_pizzas_count=pizzas_count_sold['Grande'])

                content['orderToHtml'].append({'order_id':order.order_id,'client_infos':{'name':client.name},'orderHTML':html_list_order})

        with open('izouapp/static/izouapp/data.json', 'w') as file:
            json.dump(content, file)

    return redirect(reverse('home'))

def get_date(request): # fonction qui retourne la date si elle est enregistrée dans la session sinon elle retourne la date actuelle
    if request.session.get('date_selected', None):
        return date.fromisoformat(request.session.get('date_selected'))
    return now().date().isoformat()

def datas_to_json(request): # à revoir au cas où les requetes renvoient des tableaux vides, exportent les données nécéssaire dans un fichier.json

    pizza_names = PizzaName.objects.all()
    extratoppings = ExtraTopping.objects.all()
    inventory = DailyInventory.objects.filter(date=get_date(request))
    prices = PizzaSizePrice.objects.all()
    deliveryPersons = DeliveryPerson.objects.all()

    extratoppings_ = [{'name':topping.name,'price':topping.price} for topping in extratoppings]
    deliveryPersons_ = [delivery_man.name for delivery_man in deliveryPersons]

    if len(pizza_names) == 0 or len(prices) == 0:
        pizza_ = []

    else:
        pizza_ = [{'name': n_pizza.name} for n_pizza in pizza_names]

    if len(inventory) != 0:
        if len(prices)!=0:
            inventory_ = [{'name':'Petite','pizzas_count': inventory[0].small_pizzas_count-inventory[0].sold_small_pizzas_count, 'Price':prices[0].Petite},
                           {'name':'Grande','pizzas_count': inventory[0].large_pizzas_count-inventory[0].sold_large_pizzas_count, 'Price':prices[0].Grande}]
        else:
            inventory_ = [{'name': 'Petite','pizzas_count': inventory[0].small_pizzas_count - inventory[0].sold_small_pizzas_count,'Price': 0},
                          {'name': 'Grande','pizzas_count': inventory[0].large_pizzas_count - inventory[0].sold_large_pizzas_count,'Price': 0}]

    elif len(prices)!=0:
        inventory_ = [{'name': 'Petite', 'pizzas_count': 0, 'Price':prices[0].Petite},
                            {'name': 'Grande', 'pizzas_count': 0, 'Price':prices[0].Grande}]

    else:
        inventory_ = [{'name': 'Petite', 'pizzas_count': 0, 'Price': 0},
                      {'name': 'Grande', 'pizzas_count': 0, 'Price': 0}]

    with open('izouapp/static/izouapp/data.json', 'r') as file:
        content = json.load(file)

    if pizza_==[] or len(inventory)==0 or len(prices)==0:
        content = dict()

    else:
        content['pizza'] = pizza_
        content['extratoppings'] = extratoppings_
        content['inventory'] = inventory_
        content['deliveryPersons'] = deliveryPersons_

        if 'pending_to_edit' in content.keys():
            for dict_ in content['pending_to_edit']:
                order = orders.objects.filter(order_id=dict_['order_id'])
                if not order[0].edit_requested:
                    content['pending_to_edit'].remove(dict_)

        #content.get('ordersToChart', None)
        if not 'ordersToChart' in content.keys():
            content['ordersToChart'] = {}

        if not 'orderToHtml' in content.keys():
            content['orderToHtml'] = [] # une liste de dictionnaire

        if not 'pending_to_edit' in content.keys():
            content['pending_to_edit'] = [] # une liste de dictionnaire

    with open('izouapp/static/izouapp/data.json','w') as file:
        json.dump(content, file)

def prepare_datas_to_export():
    fetched_datas = orders.objects.all()
    datas_cleaned = []
    if len(fetched_datas):
        for data in fetched_datas:
            date_formatee_fr = data.create_at.strftime("%d %B %Y")
            time_formatee = data.deliveryHour.strftime("%Hh%M")
            pizza_names = ", ".join([pizza.__str__() for pizza in data.pizzas.all()])
            surplace = 'Oui' if data.surplace else 'Non'
            delivered = 'Livré' if data.status == 'delivered' else 'Annulé'
            methode_payment = 'Chez Izoua' if data.payment_method == 'izoua' else 'Chez le livreur'
            data_row = [date_formatee_fr,surplace,data.deliveryAdress,time_formatee,data.deliveryPerson.name,
                        delivered,data.client.name,methode_payment,pizza_names,data.deliveryPrice,data.pizza_and_extratopping_price,data.total_price]

            datas_cleaned.append(data_row)

        return datas_cleaned

    return None

def create_excel_with_data(file_name, data):

    if data:
        # Construire le chemin absolu du fichier Excel dans MEDIA_ROOT
        file_path = os.path.join(settings.MEDIA_ROOT, file_name)

        # Vérifier si le fichier Excel existe
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Le fichier '{file_name}' n'existe pas dans le répertoire media.")

        # Charger le fichier Excel existant
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Insérer les données (à partir de la ligne 3)
        start_row = 3
        for row_num, row_data in enumerate(data, start=start_row):
            for col_num, value in enumerate(row_data, start=1):
                sheet.cell(row=row_num, column=col_num, value=value)

        # Sauvegarder le fichier Excel
        workbook.save(file_path)

@login_required
def download_excel(request):

    create_excel_with_data('izoua.xlsx', prepare_datas_to_export())

    # Chemin absolu du fichier Excel généré
    file_path = os.path.join(settings.MEDIA_ROOT, 'izoua.xlsx')

    # Vérifier si le fichier existe
    if not os.path.exists(file_path):
        return HttpResponse('Fichier non trouvé', status=404)

    # Ouvrir le fichier en mode binaire
    with open(file_path, 'rb') as file:
        response = HttpResponse(file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="commandes.xlsx"'
        return response

@login_required
def get_datas_to_chart(request):

    if request.method == 'POST':
        delay = int(request.POST.get('delayChart'))
        if delay<=0:
            return redirect(reverse('home'))
        datas = dict()
        days = []
        sold = []
        unsold = []

        for i in range(delay):
            daily_inventory = DailyInventory.objects.filter(
                date=date.fromisoformat(request.session['date_selected']) - timedelta(i)).first()
            if daily_inventory:
                date_to_format = date.fromisoformat(request.session['date_selected']) - timedelta(i)
                date_formatee_fr = date_to_format.strftime("%d/%m/%Y")
                days.append(date_formatee_fr)
                sold.append(daily_inventory.sold)
                unsold.append(daily_inventory.remaining)
            else:
                continue

        datas['days'] = list(reversed(days))
        datas['sold'] = list(reversed(sold))
        datas['unsold'] = list(reversed(unsold))

        with open('izouapp/static/izouapp/data.json', 'r') as file:
            content = json.load(file)

            content['ordersToChart'] = datas

        with open('izouapp/static/izouapp/data.json', 'w') as file:
            json.dump(content, file)


    return redirect(reverse('home'))

# <i class="bi bi-x-lg" onclick='deletePendingEdit(row='pending-{len_row}',order={order_to_edit})' style="background-color: green; padding: 10px; border-radius: 5px; display: inline-block;"></i>